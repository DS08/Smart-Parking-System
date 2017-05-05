##Third Party imports
from openpyxl import load_workbook

#python imports
import datetime
import sys
import os

##local imports
from opxy import parking_allowed


## vehicle can go out or in on any date
## Input Worksheet contains 5 columns:
    #(1 Column) Number Plate (Type-String)
    #(2 Column) Date (Type-String)
    #(3 Column) In-Time (Type-String)
    #(4 Column) Out-Time (Type-String)
    #(5 Column) Vehicle-In (Type-Boolean)


def main(number_plate):
    max_limit_of_parking_slots = input('Enter the maximum number of parking slots\nin the campus: ')
    filename = 'Vehicle_Registration.xlsx'
    wb = load_workbook(filename='Vehicle_Registration.xlsx')
    ws = wb.worksheets[0]
    ws.title = 'Vehicle Registration'

    rows = ws.rows

    next(rows)
    flag = 0

    count = 0

    for row in rows:
        count +=1
        if (row[0].value==number_plate):
            ## Update the entry for in-time or out-time with the help of vehicle-in
            ##  boolean field
            if (parking_allowed(max_limit_of_parking_slots)==False and row[4].value==False):
                flag = 1
                print('The vehicle is not allowed as the maximum parking slot limit is reached')
                break
            
            row[1].value = str(datetime.datetime.now().date())

            if row[4].value==True:
                row[3].value =  str(datetime.datetime.now().strftime('%I:%M %p'))
                row[4].value = False
                print('The vehicle with number plate ('+number_plate+') has been\n exited from the premises.')
            elif row[4].value == False:
                row[2].value = str(datetime.datetime.now().strftime('%I:%M %p'))
                row[3].value = ''
                row[4].value = True
                print('The vehicle with number plate ('+number_plate+') has been\n entered in the premises.')
                 
            flag = 1
            wb.save(filename)
            os.startfile(filename)
            break
            
    if (flag==0):
        if parking_allowed(max_limit_of_parking_slots):
            ws.cell(row=count+2, column=1).value = number_plate
            ws.cell(row=count+2, column=2).value = str(datetime.datetime.now().date())
            ws.cell(row=count+2, column=3).value = str(datetime.datetime.now().strftime('%I:%M %p'))
            ws.cell(row=count+2, column=4).value = ''
            ws.cell(row=count+2, column=5).value = True
            print('The vehicle with number plate ('+number_plate+') has been\n added and entered in the premises.')
            
        else:
          print('The vehicle is not allowed as the maximum parking slot limit is reached')

        wb.save(filename)
        os.startfile(filename)
    
if __name__ == "__main__":
    sys.exit(main(sys.argv[1]))
